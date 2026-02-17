"""
Celery background tasks for data ingestion from Excel files.
"""

import os
from datetime import date, datetime
from decimal import Decimal
from celery import shared_task
from django.conf import settings
import openpyxl


@shared_task(name='loans.tasks.ingest_customer_data')
def ingest_customer_data():
    """
    Ingest customer data from customer_data.xlsx.
    Expected columns (by position):
      0: Customer ID
      1: First Name
      2: Last Name
      3: Phone Number
      4: Monthly Salary
      5: Approved Limit
    """
    from customers.models import Customer

    file_path = os.path.join(settings.BASE_DIR, 'customer_data.xlsx')

    if not os.path.exists(file_path):
        return f"File not found: {file_path}"

    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # Skip header
    created_count = 0
    updated_count = 0

    for row in rows:
        if not row or row[0] is None:
            continue

        customer_id = int(row[0])
        first_name = str(row[1] or '').strip()
        last_name = str(row[2] or '').strip()
        phone_number = str(int(row[3]) if row[3] else '').strip()
        monthly_salary = Decimal(str(row[4] or 0))
        approved_limit = Decimal(str(row[5] or 0))

        customer, created = Customer.objects.update_or_create(
            customer_id=customer_id,
            defaults={
                'first_name': first_name,
                'last_name': last_name,
                'phone_number': phone_number,
                'monthly_salary': monthly_salary,
                'approved_limit': approved_limit,
                'current_debt': Decimal('0'),
            }
        )

        if created:
            created_count += 1
        else:
            updated_count += 1

    wb.close()
    return f"Customer ingestion complete: {created_count} created, {updated_count} updated."


@shared_task(name='loans.tasks.ingest_loan_data')
def ingest_loan_data():
    """
    Ingest loan data from loan_data.xlsx.
    Expected columns (by position):
      0: Customer ID
      1: Loan ID
      2: Loan Amount
      3: Tenure
      4: Interest Rate
      5: Monthly payment (EMI)
      6: EMIs paid on Time
      7: Date of Approval (start_date)
      8: End Date
    """
    from customers.models import Customer
    from loans.models import Loan

    file_path = os.path.join(settings.BASE_DIR, 'loan_data.xlsx')

    if not os.path.exists(file_path):
        return f"File not found: {file_path}"

    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # Skip header
    created_count = 0
    updated_count = 0
    skipped_count = 0
    today = date.today()

    for row in rows:
        if not row or row[0] is None:
            continue

        customer_id = int(row[0])
        loan_id = int(row[1])
        loan_amount = Decimal(str(row[2] or 0))
        tenure = int(row[3] or 0)
        interest_rate = Decimal(str(row[4] or 0))
        monthly_repayment = Decimal(str(row[5] or 0))
        emis_paid_on_time = int(row[6] or 0)

        # Parse dates — handle both datetime objects and strings
        start_date = _parse_date(row[7])
        end_date = _parse_date(row[8])

        # Determine if loan is still active
        is_active = True
        if end_date and end_date < today:
            is_active = False

        # Verify customer exists
        try:
            customer = Customer.objects.get(customer_id=customer_id)
        except Customer.DoesNotExist:
            skipped_count += 1
            continue

        loan, created = Loan.objects.update_or_create(
            loan_id=loan_id,
            defaults={
                'customer': customer,
                'loan_amount': loan_amount,
                'tenure': tenure,
                'interest_rate': interest_rate,
                'monthly_installment': monthly_repayment,
                'emis_paid_on_time': emis_paid_on_time,
                'start_date': start_date,
                'end_date': end_date,
                'is_active': is_active,
            }
        )

        if created:
            created_count += 1
        else:
            updated_count += 1

    wb.close()

    # Update current_debt for all customers based on active loans
    _update_all_customer_debts()

    # Reset PK sequences so new records get correct IDs
    _reset_pk_sequences()

    return (f"Loan ingestion complete: {created_count} created, "
            f"{updated_count} updated, {skipped_count} skipped.")


def _parse_date(value):
    """Parse a date value from xlsx — can be datetime, date, or string."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    # Try parsing string
    try:
        return datetime.strptime(str(value).strip(), '%Y-%m-%d').date()
    except (ValueError, TypeError):
        try:
            return datetime.strptime(str(value).strip(), '%d/%m/%Y').date()
        except (ValueError, TypeError):
            return None


def _update_all_customer_debts():
    """Recompute current_debt for every customer based on active loans."""
    from customers.models import Customer
    from loans.models import Loan
    from django.db.models import Sum

    customers = Customer.objects.all()
    for customer in customers:
        active_debt = Loan.objects.filter(
            customer=customer, is_active=True
        ).aggregate(
            total=Sum('loan_amount')
        )['total'] or Decimal('0')

        customer.current_debt = active_debt
        customer.save(update_fields=['current_debt'])


def _reset_pk_sequences():
    """Reset PostgreSQL auto-increment sequences after bulk data ingestion."""
    from django.db import connection

    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT setval(pg_get_serial_sequence('customers','customer_id'), "
            "COALESCE(MAX(customer_id),0)+1, false) FROM customers"
        )
        cursor.execute(
            "SELECT setval(pg_get_serial_sequence('loans','loan_id'), "
            "COALESCE(MAX(loan_id),0)+1, false) FROM loans"
        )

